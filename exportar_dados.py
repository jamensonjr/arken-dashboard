#!/usr/bin/env python3
"""
ARKEN CONSULTORIA — Exportador de Dados Financeiros
Lê ARKEN_Financeiro_2026.xlsx e gera data.json para o dashboard GitHub Pages.
"""

import json, sys, datetime, traceback
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ Execute: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

# ── CONFIGURAÇÃO ──────────────────────────────────────────────────────────────
PLANILHA_PATH = Path.home() / "Desktop" / "ARKEN CONSULTORIA" / "Financeiro" / "ARKEN_Financeiro_2026.xlsx"
SAIDA_JSON    = Path(__file__).parent / "data.json"
CAIXA_INICIAL = 48956.11
# ─────────────────────────────────────────────────────────────────────────────

MESES_ABAS  = ['JAN26','FEV26','MAR26','ABR26','MAI26','JUN26','JUL26','AGO26','SET26','OUT26','NOV26','DEZ26']
MESES_NOMES = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
DIAS_MES    = {'JAN26':31,'FEV26':28,'MAR26':31,'ABR26':30,'MAI26':31,'JUN26':30,
               'JUL26':31,'AGO26':31,'SET26':30,'OUT26':31,'NOV26':30,'DEZ26':31}
CATS_DIVIDA = ['DAS de PARCSN','PGFN-SISPAR:014629257','PGFN-SISPAR:014629258',
               'PGFN-SISPAR:014629256','Parcelamento SFR','Dívida ativa - IRPJ Fonte']
STATUS_VALIDOS_REC = {'A RECEBER','RECEBIDO','VENCIDO','CANCELADO'}

ROW_REC_REAL=17; ROW_REC_PREV=18
ROW_VAR_REAL=58; ROW_VAR_PREV=59
ROW_FIX_REAL=99; ROW_FIX_PREV=100
ROW_FLX_REAL=102; ROW_FLX_PREV=103

def v(cell):
    val = cell.value
    if isinstance(val, (int, float)): return float(val)
    return 0.0

def extrair_fluxo(wb):
    fluxo = []
    for i, aba in enumerate(MESES_ABAS):
        if aba not in wb.sheetnames: continue
        ws = wb[aba]
        ct = DIAS_MES[aba] + 3
        fluxo.append({
            "mes": MESES_NOMES[i], "aba": aba,
            "rec_real":   v(ws.cell(ROW_REC_REAL, ct)),
            "rec_prev":   v(ws.cell(ROW_REC_PREV, ct)),
            "var_real":   abs(v(ws.cell(ROW_VAR_REAL, ct))),
            "var_prev":   abs(v(ws.cell(ROW_VAR_PREV, ct))),
            "fix_real":   abs(v(ws.cell(ROW_FIX_REAL, ct))),
            "fix_prev":   abs(v(ws.cell(ROW_FIX_PREV, ct))),
            "fluxo_real": v(ws.cell(ROW_FLX_REAL, ct)),
            "fluxo_prev": v(ws.cell(ROW_FLX_PREV, ct)),
        })
    return fluxo

def extrair_detalhamento(wb):
    if 'DETALHAMENTO' not in wb.sheetnames: return []
    ws = wb['DETALHAMENTO']
    rows = []
    for row in ws.iter_rows(min_row=17, max_row=516, values_only=True):
        r = (list(row) + [None]*11)[:11]
        data, tipo, cat, _, projeto, desc, valor, _, status, comp, _ = r
        # Filtrar linhas inválidas: sem data E sem valor numérico
        if data is None and not isinstance(valor, (int, float)): continue
        # Filtrar linhas separadoras (começam com ▼ ou são cabeçalhos)
        if isinstance(data, str) and data.startswith('▼'): continue
        if isinstance(cat, str) and cat.startswith('▼'): continue
        if isinstance(data, datetime.datetime): data = data.strftime('%Y-%m-%d')
        rows.append({
            "data":    str(data or ""), "tipo": str(tipo or ""),
            "cat":     str(cat or ""), "projeto": str(projeto or ""),
            "desc":    str(desc or ""),
            "valor":   float(valor) if isinstance(valor, (int, float)) else 0.0,
            "status":  str(status or ""), "comp": str(comp or ""),
        })
    return rows

def extrair_recebiveis(wb):
    if 'RECEBÍVEIS' not in wb.sheetnames: return []
    ws = wb['RECEBÍVEIS']
    rows = []
    for row in ws.iter_rows(min_row=7, max_row=506, values_only=True):
        r = (list(row) + [None]*12)[:12]
        projeto, tipo, vt, _, parc, vp, venc, receb, status, comp, _, _ = r
        # Filtrar: sem projeto válido ou status inválido ou valor zero
        if projeto is None: continue
        if isinstance(projeto, str) and ('▼' in projeto or len(projeto.strip()) == 0): continue
        if status not in STATUS_VALIDOS_REC: continue
        if not isinstance(vp, (int, float)) or vp == 0: continue
        if isinstance(venc, datetime.datetime): venc = venc.strftime('%Y-%m-%d')
        if isinstance(receb, datetime.datetime): receb = receb.strftime('%Y-%m-%d')
        rows.append({
            "projeto":  str(projeto),
            "tipo":     str(tipo or ""),
            "val_total":float(vt) if isinstance(vt,(int,float)) else 0.0,
            "parcela":  str(parc or ""),
            "val_parc": float(vp),
            "venc":     str(venc or ""),
            "receb":    str(receb or ""),
            "status":   str(status),
            "comp":     str(comp or ""),
        })
    return rows

def extrair_projetos(wb):
    if 'MARGEM POR PROJETO' not in wb.sheetnames: return []
    ws = wb['MARGEM POR PROJETO']
    rows = []
    NOMES_INVALIDOS = {'TOTAIS','▼',''}
    for row in ws.iter_rows(min_row=4, max_row=60, values_only=True):
        r = (list(row) + [None]*9)[:5]
        nome, tipo, status, receita, custos = r
        if nome is None: continue
        nome_str = str(nome).strip()
        # Filtrar linhas de totais, separadores e vazias
        if nome_str in NOMES_INVALIDOS: continue
        if nome_str.startswith('▼'): continue
        if not isinstance(receita, (int, float)) or receita == 0: continue
        rec = float(receita)
        cus = float(custos) if isinstance(custos,(int,float)) else 0.0
        mg  = rec - cus
        rows.append({
            "nome":    nome_str,
            "tipo":    str(tipo or ""),
            "status":  str(status or ""),
            "receita": rec, "custos": cus, "margem": mg,
            "pct":     round(mg/rec*100,1) if rec else 0.0,
        })
    return rows

def extrair_dividas_detalhadas(wb):
    """Extrai saldo de cada dívida do DETALHAMENTO (status Pendente/A receber)."""
    if 'DETALHAMENTO' not in wb.sheetnames: return []
    ws = wb['DETALHAMENTO']
    pendentes = {"Pendente","A receber"}
    dividas = {}
    for cat in CATS_DIVIDA:
        dividas[cat] = 0.0
    for row in ws.iter_rows(min_row=17, max_row=516, values_only=True):
        r = (list(row) + [None]*11)[:11]
        _, _, cat, _, _, _, valor, _, status, _, _ = r
        if cat in dividas and status in pendentes and isinstance(valor, (int,float)):
            dividas[cat] += abs(float(valor))
    return [{"nome": k, "saldo": round(v, 2)} for k, v in dividas.items() if v > 0]

def calcular_kpis(fluxo, lanc, rec):
    pendentes = {"Pendente","A receber"}
    rec_real_ytd    = sum(m["rec_real"] for m in fluxo)
    custo_real_ytd  = sum(m["var_real"]+m["fix_real"] for m in fluxo)
    fluxo_real_acum = sum(m["fluxo_real"] for m in fluxo)

    a_receber = sum(r["val_parc"] for r in rec if r["status"]=="A RECEBER")
    recebido  = sum(r["val_parc"] for r in rec if r["status"]=="RECEBIDO")
    vencido   = sum(r["val_parc"] for r in rec if r["status"]=="VENCIDO")

    dividas = abs(sum(l["valor"] for l in lanc if l["cat"] in CATS_DIVIDA and l["status"] in pendentes))
    meses_c = [m for m in fluxo if m["fix_real"]+m["var_real"]>0]
    custo_fix_med = sum(m["fix_real"] for m in meses_c)/len(meses_c) if meses_c else 4591.12
    custo_var_med = sum(m["var_real"] for m in meses_c)/len(meses_c) if meses_c else 0
    custo_med = custo_fix_med + custo_var_med if meses_c else 4591.12
    caixa = CAIXA_INICIAL + fluxo_real_acum

    margem_bruta = round((rec_real_ytd - custo_real_ytd)/rec_real_ytd*100,1) if rec_real_ytd else 0
    resultado    = rec_real_ytd - custo_real_ytd

    return {
        "rec_real_ytd":      round(rec_real_ytd,2),
        "custo_real_ytd":    round(custo_real_ytd,2),
        "custo_fix_ytd":     round(sum(m["fix_real"] for m in fluxo),2),
        "custo_var_ytd":     round(sum(m["var_real"] for m in fluxo),2),
        "resultado_real":    round(resultado,2),
        "a_receber":         round(a_receber,2),
        "recebido":          round(recebido,2),
        "vencido":           round(vencido,2),
        "dividas_aberto":    round(dividas,2),
        "caixa_atual":       round(caixa,2),
        "custo_med_mensal":  round(custo_med,2),
        "custo_fix_mensal":  round(custo_fix_med,2),
        "custo_var_mensal":  round(custo_var_med,2),
        "runway_meses":      round(caixa/custo_med,1) if custo_med else 0,
        "runway_fix_meses":  round(caixa/custo_fix_med,1) if custo_fix_med else 0,
        "margem_bruta_pct":  margem_bruta,
        "comprometimento_pct": round(dividas/caixa*100,1) if caixa else 0,
        "taxa_recebimento_pct": round(recebido/(a_receber+recebido)*100,1) if (a_receber+recebido) else 0,
        "cobertura_fix": round(rec_real_ytd/sum(m["fix_real"] for m in fluxo),2) if sum(m["fix_real"] for m in fluxo) else 0,
        "break_even": round(custo_fix_med/(1-(custo_var_med/rec_real_ytd*12)) ,2) if rec_real_ytd else 0,
    }

def main():
    ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    print(f"[{ts}] Iniciando exportação ARKEN...")
    print(f"  Planilha: {PLANILHA_PATH}")

    if not PLANILHA_PATH.exists():
        print(f"  ❌ Planilha não encontrada. Ajuste PLANILHA_PATH no script.")
        return 1
    try:
        wb = openpyxl.load_workbook(str(PLANILHA_PATH), data_only=True)
        print(f"  ✅ Aberta — {len(wb.sheetnames)} abas")

        fluxo   = extrair_fluxo(wb)
        lanc    = extrair_detalhamento(wb)
        rec     = extrair_recebiveis(wb)
        proj    = extrair_projetos(wb)
        dividas = extrair_dividas_detalhadas(wb)
        kpis    = calcular_kpis(fluxo, lanc, rec)

        payload = {
            "atualizado_em": datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
            "kpis":          kpis,
            "fluxo_mensal":  fluxo,
            "recebiveis":    rec,
            "projetos":      proj,
            "dividas":       dividas,
            "lancamentos":   lanc[-50:],
        }

        SAIDA_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
        kb = SAIDA_JSON.stat().st_size / 1024
        print(f"  ✅ data.json gerado ({kb:.1f} KB)")
        print(f"  📊 {len(fluxo)} meses | {len(lanc)} lançamentos | {len(rec)} recebíveis | {len(proj)} projetos")
        print(f"  💰 Caixa: R$ {kpis['caixa_atual']:,.2f} | Runway: {kpis['runway_meses']} meses")
        return 0
    except Exception as e:
        print(f"  ❌ Erro: {e}")
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
